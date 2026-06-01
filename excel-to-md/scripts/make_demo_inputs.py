"""실증용 데모 입력 엑셀 생성기(Phase 4).

`input/` 에 멀티시트·병합셀·수식·이미지·이상치를 포함한 샘플 2~3개를 만든다.
(이 스크립트는 형상관리 대상 input/ 을 채우는 보조 도구일 뿐, 엔진의 일부는 아니다.)
"""
from __future__ import annotations

import io
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage

# 테스트 픽스처의 PNG 생성기를 재사용
import sys

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
from tests.fixtures import tiny_png_bytes  # noqa: E402


def make_sales(path: Path) -> None:
    """멀티시트 + 병합 타이틀 + 수식 + 이미지 매출표."""
    wb = Workbook()
    ws = wb.active
    ws.title = "월별매출"
    ws.merge_cells("A1:D1")
    ws["A1"] = "2026년 상반기 매출 보고"
    ws.append(["월", "매출", "비용", "이익"])
    rows = [
        ["1월", 1000, 600, 400],
        ["2월", 1200, 700, 500],
        ["3월", 1100, 650, 450],
        ["4월", 1300, 680, 620],
        ["5월", 25000, 690, 24310],  # 이상치
        ["6월", 1250, 700, 550],
    ]
    for r in rows:
        ws.append(r)
    ws["A9"] = "합계"
    ws["B9"] = "=SUM(B3:B8)"
    ws["C9"] = "=SUM(C3:C8)"
    ws.add_image(XLImage(io.BytesIO(tiny_png_bytes())), "F2")

    meta = wb.create_sheet("메타")
    meta.append(["항목", "값"])
    meta.append(["작성자", "infra-demo"])
    meta.append(["기준일", "2026-06-01"])
    wb.save(path)


def make_survey(path: Path) -> None:
    """단일 시트 설문 점수(이상치 포함, 병합 없음)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "설문응답"
    ws.append(["응답자", "만족도", "응답시간초"])
    data = [
        ["U001", 5, 30], ["U002", 4, 28], ["U003", 5, 35],
        ["U004", 3, 31], ["U005", 4, 29], ["U006", 5, 900],  # 이상치(시간)
        ["U007", 2, 27], ["U008", 4, 33],
    ]
    for r in data:
        ws.append(r)
    wb.save(path)


def make_inventory(path: Path) -> None:
    """혼합 타입(문자/수치/공백) 재고표."""
    wb = Workbook()
    ws = wb.active
    ws.title = "재고"
    ws.append(["SKU", "품목", "수량", "단가"])
    ws.append(["A-1", "노트", 120, 1500])
    ws.append(["A-2", "펜", 0, 500])
    ws.append(["A-3", "지우개", 75, ""])  # 단가 공백
    ws.append(["A-4", "파일", 40, 3000])
    wb.save(path)


if __name__ == "__main__":
    out = Path(__file__).resolve().parents[1] / "input"
    out.mkdir(parents=True, exist_ok=True)
    make_sales(out / "sales-report.xlsx")
    make_survey(out / "survey-data.xlsx")
    make_inventory(out / "inventory.xlsx")
    print("데모 입력 생성 완료:", sorted(p.name for p in out.glob("*.xlsx")))
