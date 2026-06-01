"""어댑터: XlsxReader — openpyxl 로 .xlsx 를 읽어 WorkbookData 로 변환.

추출 항목(Phase 0 실측 확정):
- 시트별 셀 값(캐시값 기준, data_only=True)
- 병합 셀 범위(merged_cells.ranges)
- 수식 셀(data_only=False 로 별도 로드)
- 임베드 이미지(openpyxl ws._images, Pillow 필요)

수식 캐시값 주의(Phase 0 근거): openpyxl 로 생성한 파일은 `<v>` 캐시가 없어
data_only=True 시 None 이 된다. Excel/LibreOffice 저장 파일은 캐시가 존재한다.
캐시가 None 이면 마크다운 렌더 단계에서 수식 문자열로 폴백한다.
"""
from __future__ import annotations

from contextlib import ExitStack
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from ..domain.errors import WorkbookReadError
from ..domain.models import (
    CellGrid,
    ImageRef,
    MergedRange,
    SheetData,
    WorkbookData,
)

# 이 리더가 처리하는 확장자
_SUPPORTED = {".xlsx", ".xlsm"}


class XlsxReader:
    """openpyxl 기반 .xlsx/.xlsm 리더(WorkbookReader 포트 구현)."""

    def supports(self, path: Path) -> bool:
        return path.suffix.lower() in _SUPPORTED

    def read(self, path: Path, assets_dir: Path | None = None) -> WorkbookData:
        try:
            wb_values = load_workbook(path, data_only=True, read_only=False)
            wb_formula = load_workbook(path, data_only=False, read_only=False)
        except Exception as exc:  # openpyxl 내부 예외를 도메인 예외로 변환
            raise WorkbookReadError(f".xlsx 읽기 실패: {path} ({exc})") from exc

        # 워크북 핸들을 ExitStack 으로 확실히 닫는다(read_only=False 라도 명시적 해제).
        with ExitStack() as stack:
            stack.callback(wb_values.close)
            stack.callback(wb_formula.close)
            sheets = tuple(
                self._read_sheet(name, wb_values[name], wb_formula[name], assets_dir)
                for name in wb_values.sheetnames
            )
            return WorkbookData(source_path=str(path), sheets=sheets)

    def _read_sheet(self, name: str, ws_val, ws_fml, assets_dir: Path | None) -> SheetData:
        """시트 1개를 SheetData 로 변환(ws_* 는 openpyxl Worksheet)."""
        grid = self._read_grid(ws_val)
        merged = self._read_merged(ws_fml)
        formulas = self._read_formulas(ws_fml)
        images = self._read_images(name, ws_fml, assets_dir)
        return SheetData(
            name=name,
            grid=grid,
            merged_ranges=merged,
            images=images,
            formulas=formulas,
        )

    @staticmethod
    def _read_grid(ws_val) -> CellGrid:
        """캐시값 기준 2차원 격자를 읽는다."""
        rows = tuple(
            tuple(cell for cell in row)
            for row in ws_val.iter_rows(values_only=True)
        )
        return CellGrid(rows=rows)

    @staticmethod
    def _read_merged(ws_fml) -> tuple[MergedRange, ...]:
        """병합 셀 범위 목록을 읽는다."""
        return tuple(
            MergedRange(
                ref=str(rng),
                min_row=rng.min_row,
                min_col=rng.min_col,
                max_row=rng.max_row,
                max_col=rng.max_col,
            )
            for rng in ws_fml.merged_cells.ranges
        )

    @staticmethod
    def _read_formulas(ws_fml) -> dict[tuple[int, int], str]:
        """수식 셀(1-based 좌표 → 수식 문자열)을 읽는다."""
        formulas: dict[tuple[int, int], str] = {}
        for row in ws_fml.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formulas[(cell.row, cell.column)] = cell.value
        return formulas

    def _read_images(self, sheet_name: str, ws_fml, assets_dir: Path | None) -> tuple[ImageRef, ...]:
        """임베드 이미지를 추출(가능 시 디스크 저장)하거나 placeholder 로 기록."""
        images: list[ImageRef] = []
        raw_images = getattr(ws_fml, "_images", [])
        for idx, img in enumerate(raw_images):
            images.append(self._extract_one_image(sheet_name, idx, img, assets_dir))
        return tuple(images)

    def _extract_one_image(
        self, sheet_name: str, idx: int, img, assets_dir: Path | None
    ) -> ImageRef:
        """이미지 1개를 추출. 실패 시 placeholder + 주석으로 graceful 처리."""
        anchor = self._anchor_to_a1(img)
        fmt = (getattr(img, "format", None) or "unknown").lower()
        if assets_dir is None:
            return ImageRef(
                anchor=anchor, image_format=fmt,
                note="이미지 감지(추출 생략: 출력 경로 미지정)",
            )
        try:
            # openpyxl 3.x private API(3.1.x 검증). 버전 변동은 아래 except 로 graceful 흡수.
            data = img._data()  # openpyxl 내부: 원본 바이트
            assets_dir.mkdir(parents=True, exist_ok=True)
            safe = "".join(c if c.isalnum() else "_" for c in sheet_name)
            fname = f"{safe}_img{idx + 1}.{fmt if fmt != 'unknown' else 'bin'}"
            (assets_dir / fname).write_bytes(data)
            return ImageRef(
                anchor=anchor, image_format=fmt,
                extracted_path=f"assets/{fname}",
            )
        except Exception as exc:  # 추출 불가 → placeholder + 주석
            return ImageRef(
                anchor=anchor, image_format=fmt,
                note=f"이미지 추출 불가: {exc}",
            )

    @staticmethod
    def _anchor_to_a1(img) -> str:
        """이미지 앵커(0-based row/col)를 A1 표기로 변환."""
        try:
            # openpyxl 3.x private API(3.1.x 검증). 버전 변동은 아래 except 로 graceful 흡수.
            marker = img.anchor._from  # type: ignore[attr-defined]
            return f"{get_column_letter(marker.col + 1)}{marker.row + 1}"
        except Exception:
            return "?"
