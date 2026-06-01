"""테스트 픽스처 — 멀티시트·병합셀·수식·이미지 포함 .xlsx 를 프로그램으로 생성.

외부 샘플 파일에 의존하지 않고 openpyxl 로 결정적으로 만들어
단위 테스트와 round-trip 검증에 사용한다.
"""
from __future__ import annotations

import io
import struct
import zlib
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage


def tiny_png_bytes() -> bytes:
    """1x1 빨강 PNG 를 순수 바이트로 생성(외부 이미지 파일 불요)."""

    def chunk(tag: bytes, data: bytes) -> bytes:
        return (
            struct.pack(">I", len(data))
            + tag
            + data
            + struct.pack(">I", zlib.crc32(tag + data) & 0xFFFFFFFF)
        )

    sig = b"\x89PNG\r\n\x1a\n"
    ihdr = chunk(b"IHDR", struct.pack(">IIBBBBB", 1, 1, 8, 2, 0, 0, 0))
    idat = chunk(b"IDAT", zlib.compress(b"\x00\xff\x00\x00"))
    iend = chunk(b"IEND", b"")
    return sig + ihdr + idat + iend


# round-trip 검증에 사용할 리터럴 데이터(수식/병합 제외 순수 값)
SALES_ROWS = [
    ["월", "매출", "비고"],
    ["1월", 100, "정상"],
    ["2월", 150, ""],
    ["3월", 120, ""],
    ["4월", 99999, "이상치"],  # 명백한 IQR/zscore 이상치 유발
]


def build_sample_workbook(path: Path) -> Path:
    """멀티시트·병합셀·수식·이미지를 포함한 샘플 워크북을 저장하고 경로 반환."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales"

    # 병합 헤더(타이틀)
    ws.merge_cells("A1:C1")
    ws["A1"] = "2026 매출표"

    # 데이터 본문(2행부터)
    for row in SALES_ROWS:
        ws.append(row)

    # 수식 셀(합계) — openpyxl 생성 파일은 캐시값이 없어 렌더 시 수식 폴백된다.
    last = len(SALES_ROWS) + 1  # 헤더(타이틀) 1 + 데이터행
    ws[f"A{last + 1}"] = "합계"
    ws[f"B{last + 1}"] = f"=SUM(B3:B{last})"

    # 임베드 이미지
    ws.add_image(XLImage(io.BytesIO(tiny_png_bytes())), "E2")

    # 두 번째 시트(메타)
    ws2 = wb.create_sheet("Meta")
    ws2.append(["키", "값"])
    ws2.append(["작성자", "fixture"])

    wb.save(path)
    return path
